import os
import re
import tkinter as tk
import time
import threading
import json

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font
from tkinter import ttk, messagebox
from datetime import datetime
from collections import defaultdict

# Define styles
fill_color_shopee = PatternFill(start_color="F7C7AC", end_color="F7C7AC", fill_type="solid")
fill_color_lazada = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
fill_color_total = PatternFill(start_color="C9C9C9", end_color="C9C9C9", fill_type="solid")
border = Border(left=Side(style='thin', color="000000"), right=Side(style='thin', color="000000"), top=Side(style='thin', color="000000"), bottom=Side(style='thin', color="000000"))
default_font = Font(name="Tahoma", size=12)  # Set default font style for rows
total_font = Font(name="Tahoma", size=12, bold=True)

valid_platforms = ["shopee", "lazada"]


def is_file_accessible(path, mode="r"):
    """Check if a file is accessible by trying to open it."""
    try:
        with open(path, mode):
            return True
    except IOError:
        return False


def get_path(year, month, day=None, report=True):
    """Return the path of the Monthly Report or Daily Bills folder for a given year, month, and day."""
    month_str = f"{month:02d}"
    base_path = os.path.join(r"C:\Users\Admin\Desktop\OrderReports", f"Year_{year}", f"Month_{month_str}")

    if report:
        return os.path.join(base_path, f"Monthly_Report_{month}_{year}.xlsx")
    elif day:
        return os.path.join(base_path, "Daily_Bills", f"Day_{day}")
    return os.path.join(base_path, "Daily_Bills")


def parse_address_zone_province_and_phone(address):
    """Extract and return the sub-district (zone), province, and phone number from an address."""
    # Remove potential phone number patterns to avoid interfering with other parsing
    phone_match = re.search(r"(Tel\.|โทร\.)?\s*(\d{3}-?\d{3}-?\d{4})", address)
    phone = phone_match.group(2).replace("-", "") if phone_match else "N/A"
    if phone != "N/A" and len(phone) == 10:
        phone = f"{phone[:3]}-{phone[3:6]}-{phone[6:]}"  # Format to XXX-XXX-XXXX

    # Remove phone number from address for cleaner parsing of zone and province
    address = re.sub(r"(Tel\.|โทร\.)?\s*\d{3}-?\d{3}-?\d{4}", "", address).strip()

    # Extract zone/sub-district
    zone = "N/A"
    if "T." in address:
        zone = address.split("T.")[1].split()[0]
    elif "ต." in address:
        zone = address.split("ต.")[1].split()[0]
    elif "ตำบล" in address:
        zone = address.split("ตำบล")[1].split()[0]

    # Extract province
    province = "N/A"
    if "จ." in address:
        province = address.split("จ.")[1].split()[0]
    elif "จังหวัด" in address:
        province = address.split("จังหวัด")[1].split()[0]
    else:
        words = address.split()
        for word in reversed(words):
            if "." not in word:
                province = word
                break

    return f"{zone} / {province}", phone


def extract_transport_service(transport_info):
    """Extract and return the transport service from the provided text."""
    return transport_info.strip() if transport_info else "N/A"


def get_last_data_row(sheet, start_row=5, column="C"):
    """Find the last row containing data in the specified column."""
    for row in range(start_row, sheet.max_row + 1):
        if sheet[f"{column}{row}"].value is None:
            return row - 1
    return sheet.max_row


def get_existing_bill_numbers(sheet, start_row=5, column="C"):
    """Return a set of existing bill numbers, truncated to the first 6 digits."""
    return {str(sheet[f"{column}{row}"].value)[:6] for row in range(start_row, sheet.max_row + 1)
            if sheet[f"{column}{row}"].value}


def apply_borders_and_format(sheet, row, start_col='A', end_col='N', font=default_font):
    """Apply borders, default font, and formatting to a row."""
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    for col in range(ord(start_col), ord(end_col) + 1):
        cell = sheet[f"{chr(col)}{row}"]
        cell.border = thin_border
        cell.font = font


def update_row_indices(sheet, start_row=5, column="A"):
    """Update row indices in the specified column, with last two rows repeating the last index."""
    last_row = get_last_data_row(sheet, start_row)
    for idx, row in enumerate(range(start_row, last_row + 1), 1):
        sheet[f"{column}{row}"].value = idx

    sheet[f"{column}{last_row + 1}"].value = f"={column}{last_row}"
    sheet[f"{column}{last_row + 2}"].value = f"={column}{last_row + 1}"


def parse_bill_filename(filename):
    """Parse the bill filename to extract a 6-character alphanumeric bill number and a non-zero box count."""
    if not filename.endswith('.xlsx'):
        return None, None

    # Extract the first 6 alphanumeric characters
    match = re.match(r"([A-Za-z0-9]{6})([0-9]{2})\.xlsx", filename)
    if match:
        bill_number, box_count = match.groups()
        if int(box_count) != 0:  # Ensure the box count is non-zero
            # Remove leading 0 if the bill number starts with it
            if bill_number.startswith('0'):
                bill_number = bill_number[1:]
            return bill_number, int(box_count)

    return None, None


def find_total_and_tax_values(sheet):
    """Identify and return total and tax values from the daily bill template."""
    total, tax, last_row = None, None, None
    for row in range(1, sheet.max_row + 1):
        if isinstance(sheet[f"F{row}"].value, (int, float)):
            total, last_row = sheet[f"F{row}"].value, row

    if last_row:
        tax = sheet[f"J{last_row + 5}"].value
    return total, tax


def convert_year(year):
    try:
        year = int(year)
        if year > 2500:  # Assuming it's a BE year if over 2500
            year -= 543
        if len(str(year)) > 4:
            raise ValueError("Year cannot have more than 4 digits.")
        return year
    except ValueError:
        messagebox.showerror("Invalid Input", "Please enter a valid 4-digit year.")
        return None


def update_days():
    try:
        year = convert_year(year_entry.get())
        month = month_map[month_var.get()]
        if not year or not month:
            return
        # Get the last day of the selected month and year
        last_day = (datetime(year, month + 1, 1) - datetime(year, month, 1)).days
        day_dropdown['values'] = list(range(1, last_day + 1))
    except Exception as e:
        messagebox.showerror("Error", str(e))


# Function to display output with badge-style updates
def write_output(message, tag="normal"):
    output_text.config(state="normal")
    output_text.insert(tk.END, message + "\n", tag)
    output_text.see(tk.END)  # Auto-scroll to the bottom
    output_text.config(state="disabled")


# Main function to process bills for the given date
def process_bills():
    year = convert_year(year_entry.get())
    month = month_map[month_var.get()]
    day = day_var.get()
    if year and month and day:
        # Start processing in a separate thread to avoid freezing the GUI
        threading.Thread(target=process_bills_thread, args=(year, month, int(day))).start()
    else:
        messagebox.showerror("Incomplete Data", "Please fill in all fields correctly.")


# Thread target for processing bills with a loading animation
def process_bills_thread(year, month, day):
    # Show loading state
    process_button.config(text="Processing...", state="disabled")
    write_output(f"Starting processing for {year}-{month:02d}-{day:02d}...")

    # Call the sample function, simulating the actual processing
    process_bills_for_day(year, month, day)

    # Restore button after processing
    process_button.config(text="Process Bills", state="normal")


def find_keyword_row(ws, keyword, column):
    """Finds the row number of the cell that contains a specific keyword in a given column."""
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=column).value == keyword:
            return row
    return None


# ---------------------------------------------------------------------------- #

def process_sheet(ws, sheet_number):
    """Processes and formats a sheet with platform-specific styling and summary rows based on the last data row."""
    last_row = get_last_data_row(ws, column="H")

    # Define red font for platform text
    platform_font_red = Font(name="Tahoma", size=12, color="FF0000")

    # Apply platform-specific formatting
    for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=9, max_col=12, values_only=False):
        transport_cell = row[2]  # Column K (Transport)
        app_cell = row[3]  # Column L (App)
        if transport_cell.value:
            match = re.search(r"/\s*(\w+)$", transport_cell.value, re.IGNORECASE)
            if match:
                platform = match.group(1).lower()
                if platform in valid_platforms:
                    app_cell.value = platform.capitalize()
                    app_cell.font = platform_font_red  # Set red font for platform text
                    app_cell.border = border
                    fill_color = fill_color_lazada if platform == "lazada" else fill_color_shopee
                    for cell in row[:3]:
                        cell.fill = fill_color
                        cell.border = border
                        cell.font = platform_font_red

    # Find dynamic rows based on labels
    shopee_row = find_keyword_row(ws, "Shopee", 5)
    lazada_row = find_keyword_row(ws, "Lazada", 5)
    grand_total_row = find_keyword_row(ws, "Grand total", 5)

    if not (shopee_row and lazada_row and grand_total_row):
        print("Required rows not found for Shopee, Lazada, or Grand Total.")
        return

    # Shopee summary formula
    ws.cell(row=shopee_row, column=6, value=f'=SUMIF(L5:L{last_row},"Shopee",F5:F{last_row})')
    ws.cell(row=shopee_row, column=7, value=f'=SUMIF(L5:L{last_row},"Shopee",G5:G{last_row})')
    ws.cell(row=shopee_row, column=8, value=f'=SUMIF(L5:L{last_row},"Shopee",H5:H{last_row})')

    # Cross-sheet reference for Shopee
    if sheet_number > 1:
        prev_sheet_name = str(sheet_number - 1)
        ws.cell(
            row=shopee_row,
            column=9,
            value=(
                f'=H{shopee_row} + SUMIF(INDIRECT("\'{prev_sheet_name}\'!E:E"), "Shopee", '
                f'INDIRECT("\'{prev_sheet_name}\'!I:I"))'
            )
        )
    else:
        ws.cell(row=shopee_row, column=9, value=f'=H{shopee_row}')

    # Lazada summary formula
    ws.cell(row=lazada_row, column=6, value=f'=SUMIF(L5:L{last_row},"Lazada",F5:F{last_row})')
    ws.cell(row=lazada_row, column=7, value=f'=SUMIF(L5:L{last_row},"Lazada",G5:G{last_row})')
    ws.cell(row=lazada_row, column=8, value=f'=SUMIF(L5:L{last_row},"Lazada",H5:H{last_row})')

    # Cross-sheet reference for Lazada
    if sheet_number > 1:
        prev_sheet_name = str(sheet_number - 1)
        ws.cell(
            row=lazada_row,
            column=9,
            value=(
                f'=H{lazada_row} + SUMIF(INDIRECT("\'{prev_sheet_name}\'!E:E"), "Lazada", '
                f'INDIRECT("\'{prev_sheet_name}\'!I:I"))'
            )
        )
    else:
        ws.cell(row=lazada_row, column=9, value=f'=H{lazada_row}')

    # Grand Total formula
    ws.cell(row=grand_total_row, column=6, value=f'=F{shopee_row}+F{lazada_row}')
    ws.cell(row=grand_total_row, column=7, value=f'=G{shopee_row}+G{lazada_row}')
    ws.cell(row=grand_total_row, column=8, value=f'=H{shopee_row}+H{lazada_row}')
    ws.cell(row=grand_total_row, column=9, value=f'=I{shopee_row}+I{lazada_row}')

    # Apply formatting to rows
    apply_borders_and_format(ws, shopee_row, font=default_font)
    apply_borders_and_format(ws, lazada_row, font=default_font)
    apply_borders_and_format(ws, grand_total_row, font=total_font)


def update_summary_formulas(sheet, start_row=5, columns=("F", "G", "H")):
    """Set up sum formulas and additional references in the last few rows for specified columns."""
    last_data_row = get_last_data_row(sheet, start_row)

    for col in columns:
        # Add sum formulas for the first summary row
        sheet[f"{col}{last_data_row + 1}"].value = f"=SUM({col}{start_row}:{col}{last_data_row})"
        # Add a reference to the first summary row in the second summary row
        sheet[f"{col}{last_data_row + 2}"].value = f"={col}{last_data_row + 1}"

        # Specifically for column 'H', add a reference two rows below the second summary row
        if col == "H":
            # Ensure there's a blank row and then add the reference formula
            sheet[f"{col}{last_data_row + 4}"].value = f"={col}{last_data_row + 2}"


# ------------------------------------------------------------------------------ #
processed_files_record = "processed_files_record.json"


def load_processed_files():
    """Load processed files record from a JSON file, creating it if it does not exist."""
    if os.path.exists(processed_files_record):
        with open(processed_files_record, "r") as file:
            return set(tuple(item) for item in json.load(file))
    else:
        # Create an empty JSON file if it doesn’t exist
        with open(processed_files_record, "w") as file:
            json.dump([], file)
        return set()


def save_processed_files(processed_files):
    """Save processed files record to a JSON file."""
    with open(processed_files_record, "w") as file:
        json.dump(list(processed_files), file)


# Load processed files from JSON on script start
processed_files = load_processed_files()


def process_bills_for_day(year, month, day):
    """Main function to process daily bills and update the monthly report."""
    report_path = get_path(year, month)
    if not is_file_accessible(report_path, mode="r+"):
        write_output(f"File {report_path} is currently open. Please close it to continue.", "warning")
        return

    daily_folder = get_path(year, month, day, report=False)

    book1_wb = load_workbook(report_path)
    sheet = book1_wb[str(day)]

    current_row = get_last_data_row(sheet) + 1 if sheet["C5"].value else 5
    bills_data = {}

    write_output("Loading workbook and checking existing bills...", "badge")

    if os.path.exists(daily_folder):

        # Dictionary to store the most recent file for each bill number
        bill_files = {}

        for filename in os.listdir(daily_folder):
            bill_number, box_count = parse_bill_filename(filename)
            if not bill_number:
                continue

            # Get the full file path and modification date
            bill_path = os.path.join(daily_folder, filename)
            mod_time = os.path.getmtime(bill_path)

            # Skip files that have already been processed in this session
            if (bill_path, mod_time) in processed_files:
                continue

            # Store the latest file for each unique bill number
            if bill_number not in bill_files or mod_time > bill_files[bill_number]['mod_time']:
                bill_files[bill_number] = {'path': bill_path, 'mod_time': mod_time, 'box_count': box_count}

        # Process only the most recent files for each unique bill number
        for bill_number, bill_info in bill_files.items():
            bill_path = bill_info['path']
            bill_wb = load_workbook(bill_path, data_only=True)
            bill_sheet = bill_wb["CashSale_th"]
            total, tax = find_total_and_tax_values(bill_sheet)
            zone_province, phone = parse_address_zone_province_and_phone(bill_sheet["D11"].value)

            if total is not None and tax is not None:
                bills_data[bill_number] = {
                    'customer_name': bill_sheet["D9"].value,
                    'zone': zone_province,
                    'box_count': bill_info['box_count'],
                    'total_value': total,
                    'tax_value': tax,
                    'transport_service': extract_transport_service(bill_sheet["D12"].value),
                    'phone': phone
                }

            bill_wb.close()
            processed_files.add((bill_path, bill_info['mod_time']))  # Track processed files in memory

    if bills_data:
        write_output("Updating rows and formulas...", "badge")
        time.sleep(2)

        for bill_number, data in bills_data.items():
            if current_row != 5:
                sheet.insert_rows(current_row)
                sheet.row_dimensions[current_row].height = 18

            sheet[f"B{current_row}"] = "PHK"
            sheet[f"C{current_row}"] = bill_number
            sheet[f"D{current_row}"] = data['customer_name']
            sheet[f"E{current_row}"] = data['zone']
            sheet[f"F{current_row}"] = data['box_count']
            sheet[f"G{current_row}"] = data['total_value']
            sheet[f"H{current_row}"] = data['tax_value']
            sheet[f"H{current_row}"].number_format = "0.00"
            sheet[f"M{current_row}"] = data['phone']
            sheet[f"K{current_row}"] = bills_data[bill_number]['transport_service']

            apply_borders_and_format(sheet, current_row)
            current_row += 1

        process_sheet(sheet, day)
        update_row_indices(sheet)
        update_summary_formulas(sheet)
        book1_wb.save(report_path)
        write_output(f"Data updated in {report_path}", "success")

        save_processed_files(processed_files)
    else:
        write_output("No new bills found.", "badge")

    book1_wb.close()


# Get current date to set as default
current_date = datetime.now()
current_year = current_date.year
current_month = current_date.month
current_day = current_date.day

# Create main window
root = tk.Tk()
root.title("Order Reports - Bill Processing")
# root.geometry("940x400")

# Get the screen width and height
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

# Set the window dimensions (800x400) and position it in the center of the screen
window_width = 940
window_height = 400

# Calculate the x and y position for centering the window
x_position = (screen_width // 2) - (window_width // 2)
y_position = (screen_height // 2) - (window_height // 2)

# Set the window geometry with the calculated position and size
root.geometry(f"{window_width}x{window_height}+{x_position}+{y_position}")

# Configure main layout
main_pane = tk.PanedWindow(root, orient=tk.HORIZONTAL, sashrelief=tk.RAISED)
main_pane.pack(fill=tk.BOTH, expand=True)

# Left panel for inputs and Process button
input_frame = tk.Frame(main_pane)
input_frame.grid_columnconfigure(1, weight=1)
main_pane.add(input_frame, width=250)

# Right panel for output display
output_frame = tk.Frame(main_pane)
output_text = tk.Text(output_frame, wrap="word", state="disabled", height=20)
output_scroll = tk.Scrollbar(output_frame, command=output_text.yview)
output_text.config(yscrollcommand=output_scroll.set)
output_text.pack(side="left", fill="both", expand=True)
output_scroll.pack(side="right", fill="y")
main_pane.add(output_frame, width=250)

# Thai month names mapped to month numbers
months_in_thai = ["มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน", "กรกฎาคม", "สิงหาคม", "กันยายน",
                  "ตุลาคม", "พฤศจิกายน", "ธันวาคม"]
month_map = {month: index + 1 for index, month in enumerate(months_in_thai)}

# Year Entry
tk.Label(input_frame, text="Year:").grid(row=0, column=0, pady=10, padx=10, sticky="w")
year_entry = tk.Entry(input_frame, width=20)
year_entry.insert(0, str(current_year))  # Set current year as default
year_entry.grid(row=0, column=1, padx=5, pady=5)

# Month Dropdown
tk.Label(input_frame, text="Month:").grid(row=1, column=0, pady=10, padx=10, sticky="w")
month_var = tk.StringVar(value=months_in_thai[current_month - 1])  # Set current month as default
month_dropdown = ttk.Combobox(input_frame, textvariable=month_var, values=months_in_thai, state="readonly", width=17)
month_dropdown.grid(row=1, column=1, padx=5, pady=5)
month_dropdown.bind("<<ComboboxSelected>>", lambda e: update_days())  # Update days when month changes

# Day Dropdown
tk.Label(input_frame, text="Day:").grid(row=2, column=0, pady=10, padx=10, sticky="w")
day_var = tk.StringVar(value=str(current_day))  # Set current day as default
day_dropdown = ttk.Combobox(input_frame, textvariable=day_var, values=[str(i) for i in range(1, 32)], state="readonly",
                            width=17)
day_dropdown.grid(row=2, column=1, padx=5, pady=5)
update_days()  # Initialize days dropdown based on current month and year

# Adjust column configuration for consistent input field alignment
input_frame.grid_columnconfigure(1, weight=1, uniform="input")

# Process Button
process_button = tk.Button(input_frame, text="Process Bills", command=process_bills)
process_button.grid(row=3, column=0, columnspan=2, pady=20)

# Text tag configurations for badge-style output
output_text.tag_configure("badge", background="#6b7280", foreground="white")
output_text.tag_configure("italic", background="#6b7280", foreground="white", font=("Arial", 10, "italic"))
output_text.tag_configure("success", background="#22c55e", foreground="white")
output_text.tag_configure("warning", background="#facc15", foreground="black")
output_text.tag_configure("error", background="#e11d48", foreground="white")

# Run the GUI
root.mainloop()

# Example usage
# process_bills_for_day(2024, 11, 7)